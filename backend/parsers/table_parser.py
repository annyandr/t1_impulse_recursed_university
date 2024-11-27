import json
import csv
import openpyxl
import xlrd
import os


def table_file_to_json(filepath, output_filepath="output.json"):
    """
    Загружает данные из табличного файла (CSV, XLS, XLSX) и сохраняет их в JSON файл.
    Тип файла определяется по расширению.

    Args:
        filepath (str): Путь к входному файлу.
        output_filepath (str, optional): Путь к выходному JSON файлу. По умолчанию "output.json".

    Returns:
        None. Печатает сообщение об успехе или ошибке.

    Raises:
        ValueError: Если указан неподдерживаемый тип файла.
    """
    try:
        file_type = os.path.splitext(filepath)[1].lower()

        if file_type not in (".csv", ".xls", ".xlsx"):
            raise ValueError(f"Неподдерживаемый тип файла: {file_type}")

        data = []

        if file_type == ".csv":
            with open(filepath, 'r', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    data.append(row)

        elif file_type == ".xls":
            wb = xlrd.open_workbook(filepath)
            sh = wb.sheet_by_index(0)
            headers = sh.row_values(0)
            for rownum in range(1, sh.nrows):
                row_data = dict(zip(headers, sh.row_values(rownum)))
                data.append(row_data)

        elif file_type == ".xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sh = wb.active
            headers = next(sh.iter_rows(values_only=True))
            for row in sh.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        with open(output_filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(data, jsonfile, indent=4, ensure_ascii=False)

        print(f"Данные из файла '{filepath}' записаны в '{output_filepath}'")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{filepath}' не найден.")
    except ValueError as e:
        print(f"Ошибка: {e}")
    except Exception as e:
        print(f"Возникла ошибка: {e}")
